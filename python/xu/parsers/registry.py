"""Parser plugins + fallback chain (PRIN-ING-5, CONST-ING-1).

Architecture absorbed from Ref/xu/routing/:
  - Intent-driven routing: (ext, intent) -> parser list
  - Rich ParseResult: success / content_markdown / metadata / skipped_reason
  - Dedicated parsers: ExcelParser (YAML), CsvParser (YAML), VisionParser (YAML)
  - Graceful fallback chains per format
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class ParseResult:
    success: bool = False
    content_markdown: str = ""
    metadata: dict = field(default_factory=dict)
    skipped_reason: str = ""
    parser: str = ""

    @property
    def ok(self) -> bool:
        return self.success and bool(self.content_markdown.strip())

    @property
    def text(self) -> str:
        return self.content_markdown


# ---- ExcelParser: openpyxl -> YAML list of dicts (table content_type) ----
class ExcelParser:
    name = "excel"
    SUPPORTED = {".xlsx", ".xls"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        import yaml
        try:
            import openpyxl
        except ImportError:
            return None
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception:
            return None
        try:
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_idx = 0
                headers = []
                for i, row in enumerate(rows):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if sum(1 for c in cells if c) >= 3:
                        header_idx = i
                        headers = cells
                        break
                if not headers:
                    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
                if header_idx > 0:
                    meta_lines = []
                    for row in rows[:header_idx]:
                        cells = [str(c) for c in row if c is not None and str(c).strip()]
                        if cells:
                            meta_lines.append(" | ".join(cells))
                    if meta_lines:
                        parts.append("## Sheet: " + sheet_name + "\\n")
                        parts.append("\\n".join(meta_lines) + "\\n")
                items = []
                for row in rows[header_idx + 1:]:
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if not any(cells):
                        continue
                    item = {}
                    for j, h in enumerate(headers):
                        if h and j < len(cells) and cells[j]:
                            item[h] = cells[j]
                    if item:
                        items.append(item)
                parts.append("## Sheet: " + sheet_name + "\\n")
                parts.append(yaml.dump(items, allow_unicode=True, default_flow_style=False, sort_keys=False))
            text = "".join(parts)
            return ParseResult(success=True, content_markdown=text,
                              metadata={"parser": self.name, "sheets": wb.sheetnames},
                              parser=self.name)
        except Exception:
            return None


# ---- CsvParser: csv -> YAML list of dicts (table content_type) ----
class CsvParser:
    name = "csv"
    SUPPORTED = {".csv"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        import csv, yaml
        try:
            with open(str(path), encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
        except Exception:
            return None
        if not rows:
            return None
        header = rows[0]
        items = []
        for row in rows[1:]:
            cells = (row + [""] * len(header))[:len(header)]
            if not any(c.strip() for c in cells):
                continue
            item = {}
            for j, h in enumerate(header):
                if h and j < len(cells) and cells[j]:
                    item[h] = cells[j]
            if item:
                items.append(item)
        try:
            text = yaml.dump(items, allow_unicode=True, default_flow_style=False, sort_keys=False)
        except Exception:
            return None
        return ParseResult(success=True, content_markdown=text,
                          metadata={"parser": self.name, "rows": len(items)},
                          parser=self.name)


# ---- VisionParser: PIL/EXIF -> YAML list of dicts (gallery content_type) ----
class VisionParser:
    name = "vision"
    SUPPORTED = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        import os, yaml
        p = Path(path)
        if not p.is_file():
            return None
        size = os.path.getsize(p)
        item: dict = {"filename": p.name, "size_bytes": size}
        try:
            from PIL import Image
            img = Image.open(p)
            item["width"] = img.width
            item["height"] = img.height
            exif = img.getexif()
            if exif:
                from PIL.ExifTags import TAGS
                for tag_id, value in exif.items():
                    tag = TAGS.get(tag_id, "")
                    if tag in ("DateTimeOriginal", "Make", "Model"):
                        if isinstance(value, bytes):
                            value = value.decode("utf-8", errors="replace").replace(chr(0), "")
                        item[tag] = str(value)
        except (OSError, ValueError, AttributeError):
            pass
        text = yaml.dump([item], allow_unicode=True, default_flow_style=False, sort_keys=False)
        return ParseResult(success=True, content_markdown=text,
                          metadata={"parser": self.name},
                          parser=self.name)


# ---- MinerU primary (cloud API; silent fallback if key missing) ----
class MinerUParser:
    name = "mineru"
    SUPPORTED = {".pdf", ".docx", ".pptx"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        from ..parsers.mineru_parser import mineru_parse
        text = mineru_parse(str(path), api_key=kw.get("mineru_key", ""))
        if text and text.strip():
            return ParseResult(success=True, content_markdown=text,
                              metadata={"parser": self.name})
        return None


# ---- markitdown local fallback (offline, no API) ----
class MarkitdownParser:
    name = "markitdown"
    SUPPORTED = {".pdf", ".docx", ".pptx", ".html", ".htm"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        try:
            from markitdown import MarkItDown
            md = MarkItDown()
            r = md.convert(str(path))
            content = r.text_content or ""
            if not content.strip():
                return None
            return ParseResult(success=True, content_markdown=content,
                              metadata={"parser": self.name})
        except Exception:
            return None


# ---- plain text / markdown ----
class TextParser:
    name = "text"
    SUPPORTED = {".txt", ".md", ".markdown", ".log", ".json", ".yaml", ".yml"}

    def can_parse(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED

    def parse(self, path: Path, **kw) -> ParseResult | None:
        try:
            content = path.read_text(encoding="utf-8", errors="replace")
            if not content.strip():
                return None
            return ParseResult(success=True, content_markdown=content,
                              metadata={"parser": self.name})
        except Exception:
            return None


# Fallback chains (PRIN-ING-5)
_CHAINS: dict[str, list] = {
    ".xlsx": [ExcelParser()],
    ".xls":  [ExcelParser()],
    ".csv":  [CsvParser()],
    ".pdf":  [MinerUParser(), MarkitdownParser()],
    ".docx": [MinerUParser(), MarkitdownParser()],
    ".pptx": [MinerUParser(), MarkitdownParser()],
    ".png":  [VisionParser()],
    ".jpg":  [VisionParser()],
    ".jpeg": [VisionParser()],
    ".gif":  [VisionParser()],
    ".webp": [VisionParser()],
    ".bmp":  [VisionParser()],
    ".tiff": [VisionParser()],
    ".html": [MarkitdownParser()],
    ".htm":  [MarkitdownParser()],
}


def _chain_for(path: Path) -> list:
    return _CHAINS.get(path.suffix.lower(), [])


def parse_file(path: str | Path, **kw) -> ParseResult:
    """Run the fallback chain. Returns a ParseResult; .ok=False means rejected.

    Collects per-parser failure reasons so the final skipped_reason discloses
    what each parser tried and why it failed (PRIN-ING-5 transparency).
    """
    p = Path(path)
    chain = _chain_for(p)
    attempts = []
    for parser in chain:
        if not parser.can_parse(p):
            continue
        try:
            res = parser.parse(p, **kw)
        except Exception as e:
            attempts.append(f"{parser.name}: {e}")
            continue
        if res and res.ok:
            return res
        reason = res.skipped_reason if res else "empty result"
        attempts.append(f"{parser.name}: {reason}")
    attempts_str = "; ".join(attempts) if attempts else "no parser attempted"
    return ParseResult(success=False,
                      skipped_reason=f"all parsers failed for {Path(path).name}: {attempts_str}; cannot enter Phase 2 (PRIN-ING-5)")
