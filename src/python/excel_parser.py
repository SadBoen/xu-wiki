"""excel_parser — thin wrapper around openpyxl for Rust via pyo3.

No business logic. Pure library interface.
Supports: .xlsx .xls
Returns YAML-formatted content (table content_type).
"""
from __future__ import annotations

import csv
import os
import yaml
from pathlib import Path


SUPPORTED = {".xlsx", ".xls", ".csv"}


def can_parse(path: str) -> bool:
    """Return True if the file extension is supported."""
    return Path(path).suffix.lower() in SUPPORTED


def parse_excel(path: str) -> str:
    """Parse an Excel workbook and return all sheets as YAML.

    Returns empty string on any error.
    """
    p = Path(path)
    if not p.is_file():
        return ""

    ext = p.suffix.lower()
    if ext == ".csv":
        return _parse_csv(p)
    return _parse_xlsx(p)


def _parse_xlsx(p: Path) -> str:
    """Parse .xlsx / .xls via openpyxl, return YAML list-of-dicts."""
    try:
        import openpyxl
    except ImportError:
        return ""

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception:
        return ""

    try:
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row (first row with >= 3 non-empty cells)
            header_idx = 0
            headers: list[str] = []
            for i, row in enumerate(rows):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if sum(1 for c in cells if c) >= 3:
                    header_idx = i
                    headers = cells
                    break
            if not headers:
                headers = [str(c).strip() if c is not None else "" for c in rows[0]]

            # Meta rows above header
            if header_idx > 0:
                meta_lines = []
                for row in rows[:header_idx]:
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        meta_lines.append(" | ".join(cells))
                if meta_lines:
                    parts.append(f"## Sheet: {sheet_name}\n")
                    parts.append("\n".join(meta_lines) + "\n")

            # Data rows
            items: list[dict[str, str]] = []
            for row in rows[header_idx + 1 :]:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if not any(cells):
                    continue
                item = {}
                for j, h in enumerate(headers):
                    if h and j < len(cells) and cells[j]:
                        item[h] = cells[j]
                if item:
                    items.append(item)

            parts.append(f"## Sheet: {sheet_name}\n")
            parts.append(
                yaml.dump(
                    items,
                    allow_unicode=True,
                    default_flow_style=False,
                    sort_keys=False,
                )
            )

        return "".join(parts)
    except Exception:
        return ""


def _parse_csv(p: Path) -> str:
    """Parse a .csv file, return YAML list-of-dicts."""
    try:
        with open(str(p), encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
    except Exception:
        return ""

    if not rows:
        return ""

    header = rows[0]
    items: list[dict[str, str]] = []
    for row in rows[1:]:
        cells = (row + [""] * len(header))[: len(header)]
        if not any(c.strip() for c in cells):
            continue
        item = {}
        for j, h in enumerate(header):
            if h and j < len(cells) and cells[j]:
                item[h] = cells[j]
        if item:
            items.append(item)

    try:
        return yaml.dump(
            items,
            allow_unicode=True,
            default_flow_style=False,
            sort_keys=False,
        )
    except Exception:
        return ""


def get_excel_as_markdown(path: str) -> str:
    """Alias for parse_excel for backward compatibility."""
    return parse_excel(path)
